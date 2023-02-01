"""
    Esse parte do projeto irá transformar a malha conforme a ferramenta
    de dimensionamento do capacity planning após isso estará pronta para
    se utilizada para o dimensionamento de demanda.

"""
from openpyxl import load_workbook
import datetime

Base = 'VCP'
time_null = '00:00'

"""Nomear colunas conforme a ferramenta de dimensionamento do capacity"""


class Sir:

    def __init__(self):
        """Variáveis que receberam o valor por condição em looping"""

        self.dept_sta = ''
        self.arvl_sta = ''
        self.dept_day = ''
        self.arvl_day = ''
        self.dept_time = ''
        self.arvl_time = ''
        self.sub_fleet = ''
        self.flight_number = ''
        self.trilho_set = ''
        self.svc_type = ''
        self.week_day = ''
        self.identity = ''
        """Pequenos Parametros"""

        self.base = Base
        self.tempo = time_null
        self.linha = 2
        self.coluna = 1
        self.escrever_coluna = 1
        self.escrever_linha = 2

        """Carrega a planilha tratada pelo pandas assim atibuindos respectivos valores das colunas"""

        self.malha = load_workbook(f'SIR - MALHA {datetime.date.today()}.xlsx')
        self.sir = self.malha.active
        self.deptsta = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.arlvsta = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.deptday = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.arvlday = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.depttime = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.arvltime = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.subfleet = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.flightnumberdept = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.flightnumberarvl = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.trilho = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.svctype = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.weekday = self.sir.cell(column=self.coluna, row=self.linha).value
        self.coluna += 1
        self.identity = self.sir.cell(column=self.coluna, row=self.linha).value

        """Cria e nova planilha e escreve o cabeçalho"""

        self.planilha = self.malha.create_sheet('MALHA', 1)
        self.planilha['A1'] = "Dept Sta"
        self.planilha['B1'] = "Arvl Sta"
        self.planilha['C1'] = "Dept Day"
        self.planilha['D1'] = "Arvl Day"
        self.planilha['E1'] = "Dept Time"
        self.planilha['F1'] = "Arvl Time"
        self.planilha['G1'] = "Subfleet"
        self.planilha['H1'] = "FlightNumber"
        self.planilha['I1'] = "Trilho"
        self.planilha['J1'] = "Svc Type"
        self.planilha['K1'] = "Week Day"
        self.planilha['L1'] = "ID"

        """Extraindo dados da primeira planilha e transformando para o formato Capacity Planning"""


class EscreverMalha(Sir):

    def __init__(self):
        super().__init__()

        """Escreve os dados na planilha conforme a coluna e a linha"""

        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.dept_sta
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.arvl_sta
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.dept_day
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.arvl_day
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.dept_time
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.arvl_time
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.sub_fleet
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.flight_number
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.trilho_set
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.svc_type
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.week_day
        self.escrever_coluna += 1
        self.planilha.cell(column=self.escrever_coluna, row=self.escrever_linha).value = self.identity
        self.escrever_coluna -= 11
        self.escrever_linha += 1

        """Looping que define o valor das variáveis"""

        if self.identity == 43 and self.deptsta != Base:
            self.dept_sta = self.deptsta
            self.arvl_sta = Base
            self.dept_day = self.deptday
            self.arvl_day = self.arvlday
            self.dept_time = self.tempo
            self.arvl_time = f'{self.arvlday} {self.arvltime}'
            self.sub_fleet = self.subfleet
            self.flight_number = self.flightnumberarvl
            self.trilho_set = self.trilho
            self.svc_type = self.svctype
            self.week_day = self.weekday
            self.identity = self.identity
            EscreverMalha()

        elif self.identity == 43 and self.deptsta == Base:

            self.dept_sta = Base
            self.arvl_sta = self.arlvsta
            self.dept_day = self.deptday
            self.arvl_day = self.arvlday
            self.dept_time = f'{self.dept_day} {self.dept_time}'
            self.arvl_time = self.tempo
            self.sub_fleet = self.subfleet
            self.flight_number = self.flightnumberdept
            self.trilho = self.trilho
            self.svc_type = self.svctype
            self.week_day = self.weekday
            self.identity = self.identity
            EscreverMalha()

        elif self.identity == 44 and self.deptsta != Base:

            self.dept_sta = self.deptsta
            self.arvl_sta = Base
            self.dept_day = self.deptday
            self.arvl_day = self.arvlday
            self.dept_time = self.tempo
            self.arvl_time = f'{self.arvlday} {self.arvltime}'
            self.sub_fleet = self.subfleet
            self.flight_number = self.flightnumberarvl
            self.trilho = self.trilho
            self.svc_type = self.svctype
            self.week_day = self.weekday
            self.identity = self.identity
            EscreverMalha()

        elif self.identity == 44 and self.deptsta == Base:

            self.dept_sta = Base
            self.arvl_sta = self.arlvsta
            self.dept_day = self.deptday
            self.arvl_day = self.arvlday
            self.dept_time = f'{self.dept_day} {self.dept_time}'
            self.arvl_time = self.tempo
            self.sub_fleet = self.subfleet
            self.flight_number = self.flightnumberdept
            self.trilho = self.trilho
            self.svc_type = self.svctype
            self.week_day = self.weekday
            self.identity = self.identity
            EscreverMalha()

        elif self.identity == 45 and self.deptsta != Base:

            self.dept_sta = self.deptsta
            self.arvl_sta = Base
            self.dept_day = self.deptday
            self.arvl_day = self.arvlday
            self.dept_time = self.tempo
            self.arvl_time = f'{self.arvlday} {self.arvltime}'
            self.sub_fleet = self.subfleet
            self.flight_number = self.flightnumberarvl
            self.trilho = self.trilho
            self.svc_type = self.svctype
            self.week_day = self.weekday
            self.identity = self.identity
            EscreverMalha()

        elif self.identity == 46 and self.deptsta != Base and self.arlvsta != Base:
            """Escreve a Saida"""
            self.dept_sta = self.deptsta
            self.arvl_sta = Base
            self.dept_day = self.deptday
            self.arvl_day = self.arvlday
            self.dept_time = self.tempo
            self.arvl_time = f'{self.arvlday} {self.arvltime}'
            self.sub_fleet = self.subfleet
            self.flight_number = self.flightnumberarvl
            self.trilho = self.trilho
            self.svc_type = self.svctype
            self.week_day = self.weekday
            self.identity = self.identity
            EscreverMalha()

            """Escreve a Saida"""
            self.dept_sta = Base
            self.arvl_sta = self.arlvsta
            self.dept_day = self.deptday
            self.arvl_day = self.arvlday
            self.dept_time = f'{self.dept_day} {self.dept_time}'
            self.arvl_time = self.tempo
            self.sub_fleet = self.subfleet
            self.flight_number = self.flightnumberdept
            self.trilho = self.trilho
            self.svc_type = self.svctype
            self.week_day = self.weekday
            self.identity = self.identity
            EscreverMalha()
            self.malha.save(filename=f'SIR - MALHA {datetime.date.today()}.xlsx')
